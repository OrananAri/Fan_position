import openpyxl as xl
import numpy as np

workbook = xl.load_workbook('position.xlsx')
sheet = workbook['Sheet1']

fan_table=np.array([[0.000000000] * 7 for i in range(2, sheet.max_row + 1)])
np.set_printoptions(precision=4)
# print(fan_table)

print(f"# of fan = {sheet.max_row-2+1}")
# print(range(2,sheet.max_row))
# print(sheet.cell(2,1).value)

#นำค่า x จาก excel เข้าตาราง matrix
r=0
for i in range(2, sheet.max_row + 1):
        fan_table[r][0]=sheet.cell(i,1).value
        # print(f"{r} {i} {sheet.cell(i,1).value}")
        r += 1

#นำค่า y จาก excel เข้าตาราง matrix
r=0
for i in range(2, sheet.max_row + 1):
        fan_table[r][1]=sheet.cell(i,2).value
        r += 1

#นำค่า z จาก excel เข้าตาราง matrix
r=0
for i in range(2, sheet.max_row + 1):
        fan_table[r][2]=sheet.cell(i,3).value
        r += 1

#นำค่า direction จาก excel เข้าตาราง matrix
#1=left, 2=right, 3=up, 4=down
r=0
for i in range(2, sheet.max_row + 1):
        fan_table[r][3]=sheet.cell(i,4).value
        r += 1

#นำค่า u จาก excel เข้าตาราง matrix
r=0
for i in range(2, sheet.max_row + 1):
        fan_table[r][4]=sheet.cell(i,5).value
        r += 1

# #นำค่า v จาก excel เข้าตาราง matrix
r=0
for i in range(2, sheet.max_row + 1):
        fan_table[r][5]=sheet.cell(i,6).value
        r += 1

#นำค่า w จาก excel เข้าตาราง matrix
r=0
for i in range(2, sheet.max_row + 1):
        fan_table[r][6]=sheet.cell(i,7).value
        r += 1

#print(fan_table)

fan_length=2
fan_width=0.5
fan_height=0.5


x_min=np.array([[0.000000000] for i in range(1, sheet.max_row + 1)])
np.set_printoptions(precision=4)
# print(f"{x_min}")
x_max=np.array([[0.000000000] for i in range(1, sheet.max_row + 1)])
np.set_printoptions(precision=4)
y_min=np.array([[0.000000000] for i in range(1, sheet.max_row + 1)])
np.set_printoptions(precision=4)
y_max=np.array([[0.000000000] for i in range(1, sheet.max_row + 1)])
np.set_printoptions(precision=4)
z_min=np.array([[0.000000000] for i in range(1, sheet.max_row + 1)])
np.set_printoptions(precision=4)
z_max=np.array([[0.000000000] for i in range(1, sheet.max_row + 1)])
np.set_printoptions(precision=4)
u=np.array([[0.000000000] for i in range(1, sheet.max_row + 1)])
np.set_printoptions(precision=4)
v=np.array([[0.000000000] for i in range(1, sheet.max_row + 1)])
np.set_printoptions(precision=4)
w=np.array([[0.000000000] for i in range(1, sheet.max_row + 1)])
np.set_printoptions(precision=4)
#
#
b=1
volumeflow=1.630000
##คำนวณตัวแปร
# # with open('NewFormat_output.fds', 'w') as OutputFile_NewFormat:

for r in range(0, sheet.max_row - 2 + 1):
        # print(r)
        # direction = left
      if fan_table[r][3]==1:
            x_min[b] = fan_table[r][0] - (fan_length / 2)
            x_max[b] = fan_table[r][0] + (fan_length / 2)
            y_min[b] = fan_table[r][1] - (fan_width / 2)
            y_max[b] = fan_table[r][1] + (fan_width / 2)
            z_min[b] = fan_table[r][2] - (fan_height / 2)
            z_max[b] = fan_table[r][2] + (fan_height / 2)

            print(f"&OBST ID='B{b}', XB={x_min[b]},{x_max[b]},{y_min[b]},{y_max[b]},{z_min[b]},{z_max[b]}, SURF_ID='INERT'/")

            #without uvw
            if fan_table[r][4] == 0 and fan_table[r][5] == 0:
                print(f"&VENT ID = 'V{b}01', SURF_ID = 'HVAC', XB = {x_max[b]}, {x_max[b]}, {y_min[b]}, {y_max[b]}, {z_min[b]}, {z_max[b]}, RADIUS = 0.225, XYZ = {x_max[b]}, {fan_table[r][1]:.4f}, {fan_table[r][2]:.4f}/")
                print(f"&VENT ID = 'V{b}02', SURF_ID = 'HVAC', XB = {x_min[b]}, {x_min[b]}, {y_min[b]}, {y_max[b]}, {z_min[b]}, {z_max[b]}, RADIUS = 0.225, XYZ = {x_min[b]}, {fan_table[r][1]:.4f}, {fan_table[r][2]:.4f}/")
                print(f"&HVAC ID = 'N{b}01', TYPE_ID = 'NODE', DUCT_ID = 'D{b}', VENT_ID = 'V{b}01'/")
                print(f"&HVAC ID = 'N{b}02', TYPE_ID = 'NODE', DUCT_ID = 'D{b}', VENT_ID = 'V{b}02'/")
                print(f"&HVAC ID = 'D{b}', TYPE_ID = 'DUCT', DIAMETER = 0.45, VOLUME_FLOW = {volumeflow:.4f}, NODE_ID = 'N{b}01', 'N{b}02', ROUGHNESS = 1.0E-3, LENGTH = {fan_length:.4f}/")
                b += 1

            #contain uvw
            else:
                print(f"&VENT ID = 'V{b}01', SURF_ID = 'HVAC', XB = {x_max[b]}, {x_max[b]}, {y_min[b]}, {y_max[b]}, {z_min[b]}, {z_max[b]}, RADIUS = 0.225, XYZ = {x_max[b]}, {fan_table[r][1]:.4f}, {fan_table[r][2]:.4f}/")
                print(f"&VENT ID = 'V{b}02', SURF_ID = 'HVAC', XB = {x_min[b]}, {x_min[b]}, {y_min[b]}, {y_max[b]}, {z_min[b]}, {z_max[b]}, RADIUS = 0.225, XYZ = {x_min[b]}, {fan_table[r][1]:.4f}, {fan_table[r][2]:.4f}, UVW={fan_table[r][4]:.4f},{fan_table[r][5]:.4f},{fan_table[r][6]:.4f}/")
                print(f"&HVAC ID = 'N{b}01', TYPE_ID = 'NODE', DUCT_ID = 'D{b}', VENT_ID = 'V{b}01'/")
                print(f"&HVAC ID = 'N{b}02', TYPE_ID = 'NODE', DUCT_ID = 'D{b}', VENT_ID = 'V{b}02'/")
                print(f"&HVAC ID = 'D{b}', TYPE_ID = 'DUCT', DIAMETER = 0.45, VOLUME_FLOW = {volumeflow:.4f}, NODE_ID = 'N{b}01', 'N{b}02', ROUGHNESS = 1.0E-3, LENGTH = {fan_length:.4f}/")
                b += 1

        # direction = right
      elif fan_table[r][3]==2:
          x_min[b] = fan_table[r][0] - (fan_length / 2)
          x_max[b] = fan_table[r][0] + (fan_length / 2)
          y_min[b] = fan_table[r][1] - (fan_width / 2)
          y_max[b] = fan_table[r][1] + (fan_width / 2)
          z_min[b] = fan_table[r][2] - (fan_height / 2)
          z_max[b] = fan_table[r][2] + (fan_height / 2)

          print(f"&OBST ID='B{b}', XB={x_min[b]},{x_max[b]},{y_min[b]},{y_max[b]},{z_min[b]},{z_max[b]}, SURF_ID='INERT'/")

          # without uvw
          if fan_table[r][4] == 0 and fan_table[r][5] == 0:
              print(f"&VENT ID = 'V{b}01', SURF_ID = 'HVAC', XB = {x_min[b]}, {x_min[b]}, {y_min[b]}, {y_max[b]}, {z_min[b]}, {z_max[b]}, RADIUS = 0.225, XYZ = {x_min[b]}, {fan_table[r][1]:.4f}, {fan_table[r][2]:.4f}/")
              print(f"&VENT ID = 'V{b}02', SURF_ID = 'HVAC', XB = {x_max[b]}, {x_max[b]}, {y_min[b]}, {y_max[b]}, {z_min[b]}, {z_max[b]}, RADIUS = 0.225, XYZ = {x_max[b]}, {fan_table[r][1]:.4f}, {fan_table[r][2]:.4f}/")
              print(f"&HVAC ID = 'N{b}01', TYPE_ID = 'NODE', DUCT_ID = 'D{b}', VENT_ID = 'V{b}01'/")
              print(f"&HVAC ID = 'N{b}02', TYPE_ID = 'NODE', DUCT_ID = 'D{b}', VENT_ID = 'V{b}02'/")
              print(f"&HVAC ID = 'D{b}', TYPE_ID = 'DUCT', DIAMETER = 0.45, VOLUME_FLOW = {volumeflow:.4f}, NODE_ID = 'N{b}01', 'N{b}02', ROUGHNESS = 1.0E-3, LENGTH = {fan_length:.4f}/")
              b += 1

          # contain uvw
          else:
              print(f"&VENT ID = 'V{b}01', SURF_ID = 'HVAC', XB = {x_min[b]}, {x_min[b]}, {y_min[b]}, {y_max[b]}, {z_min[b]}, {z_max[b]}, RADIUS = 0.225, XYZ = {x_min[b]}, {fan_table[r][1]:.4f}, {fan_table[r][2]:.4f}/")
              print(f"&VENT ID = 'V{b}02', SURF_ID = 'HVAC', XB = {x_max[b]}, {x_max[b]}, {y_min[b]}, {y_max[b]}, {z_min[b]}, {z_max[b]}, RADIUS = 0.225, XYZ = {x_max[b]}, {fan_table[r][1]:.4f}, {fan_table[r][2]:.4f}, UVW={fan_table[r][4]:.4f},{fan_table[r][5]:.4f},{fan_table[r][6]:.4f}/")
              print(f"&HVAC ID = 'N{b}01', TYPE_ID = 'NODE', DUCT_ID = 'D{b}', VENT_ID = 'V{b}01'/")
              print(f"&HVAC ID = 'N{b}02', TYPE_ID = 'NODE', DUCT_ID = 'D{b}', VENT_ID = 'V{b}02'/")
              print(f"&HVAC ID = 'D{b}', TYPE_ID = 'DUCT', DIAMETER = 0.45, VOLUME_FLOW = {volumeflow:.4f}, NODE_ID = 'N{b}01', 'N{b}02', ROUGHNESS = 1.0E-3, LENGTH = {fan_length:.4f}/")
              b += 1

      # direction = up
      elif fan_table[r][3]==3:
          x_min[b] = fan_table[r][0] - (fan_width / 2)
          x_max[b] = fan_table[r][0] + (fan_width / 2)
          y_min[b] = fan_table[r][1] - (fan_length / 2)
          y_max[b] = fan_table[r][1] + (fan_length / 2)
          z_min[b] = fan_table[r][2] - (fan_height / 2)
          z_max[b] = fan_table[r][2] + (fan_height / 2)
          print(f"&OBST ID='B{b}', XB={x_min[b]},{x_max[b]},{y_min[b]},{y_max[b]},{z_min[b]},{z_max[b]}, SURF_ID='INERT'/")

          # without uvw
          if fan_table[r][4] == 0 and fan_table[r][5] == 0:
              print(f"&VENT ID = 'V{b}01', SURF_ID = 'HVAC', XB = {x_min[b]}, {x_max[b]}, {y_min[b]}, {y_min[b]}, {z_min[b]}, {z_max[b]}, RADIUS = 0.225, XYZ = {fan_table[r][0]:.4f}, {y_min[b]}, {fan_table[r][2]:.4f}/")
              print(f"&VENT ID = 'V{b}02', SURF_ID = 'HVAC', XB = {x_min[b]}, {x_max[b]}, {y_max[b]}, {y_max[b]}, {z_min[b]}, {z_max[b]}, RADIUS = 0.225, XYZ = {fan_table[r][0]:.4f}, {y_max[b]}, {fan_table[r][2]:.4f}/")
              print(f"&HVAC ID = 'N{b}01', TYPE_ID = 'NODE', DUCT_ID = 'D{b}', VENT_ID = 'V{b}01'/")
              print(f"&HVAC ID = 'N{b}02', TYPE_ID = 'NODE', DUCT_ID = 'D{b}', VENT_ID = 'V{b}02'/")
              print(f"&HVAC ID = 'D{b}', TYPE_ID = 'DUCT', DIAMETER = 0.45, VOLUME_FLOW = {volumeflow:.4f}, NODE_ID = 'N{b}01', 'N{b}02', ROUGHNESS = 1.0E-3, LENGTH = {fan_length:.4f}/")
              b += 1

          # contain uvw
          else:
              print(f"&VENT ID = 'V{b}01', SURF_ID = 'HVAC', XB = {x_min[b]}, {x_max[b]}, {y_min[b]}, {y_min[b]}, {z_min[b]}, {z_max[b]}, RADIUS = 0.225, XYZ = {fan_table[r][0]:.4f}, {y_min[b]}, {fan_table[r][2]:.4f}/")
              print(f"&VENT ID = 'V{b}02', SURF_ID = 'HVAC', XB = {x_min[b]}, {x_max[b]}, {y_max[b]}, {y_max[b]}, {z_min[b]}, {z_max[b]}, RADIUS = 0.225, XYZ = {fan_table[r][0]:.4f}, {y_max[b]}, {fan_table[r][2]:.4f}, UVW={fan_table[r][4]:.4f},{fan_table[r][5]:.4f},{fan_table[r][6]:.4f}/")
              print(f"&HVAC ID = 'N{b}01', TYPE_ID = 'NODE', DUCT_ID = 'D{b}', VENT_ID = 'V{b}01'/")
              print(f"&HVAC ID = 'N{b}02', TYPE_ID = 'NODE', DUCT_ID = 'D{b}', VENT_ID = 'V{b}02'/")
              print(f"&HVAC ID = 'D{b}', TYPE_ID = 'DUCT', DIAMETER = 0.45, VOLUME_FLOW = {volumeflow:.4f}, NODE_ID = 'N{b}01', 'N{b}02', ROUGHNESS = 1.0E-3, LENGTH = {fan_length:.4f}/")
              b += 1

      # direction = down
      elif fan_table[r][3]==4:
          x_min[b] = fan_table[r][0] - (fan_width / 2)
          x_max[b] = fan_table[r][0] + (fan_width / 2)
          y_min[b] = fan_table[r][1] - (fan_length / 2)
          y_max[b] = fan_table[r][1] + (fan_length / 2)
          z_min[b] = fan_table[r][2] - (fan_height / 2)
          z_max[b] = fan_table[r][2] + (fan_height / 2)
          print(f"&OBST ID='B{b}', XB={x_min[b]},{x_max[b]},{y_min[b]},{y_max[b]},{z_min[b]},{z_max[b]}, SURF_ID='INERT'/")

          # without uvw
          if fan_table[r][4] == 0 and fan_table[r][5] == 0:
              print(f"&VENT ID = 'V{b}01', SURF_ID = 'HVAC', XB = {x_min[b]}, {x_max[b]}, {y_max[b]}, {y_max[b]}, {z_min[b]}, {z_max[b]}, RADIUS = 0.225, XYZ = {fan_table[r][0]:.4f}, {y_max[b]}, {fan_table[r][2]:.4f}/")
              print(f"&VENT ID = 'V{b}02', SURF_ID = 'HVAC', XB = {x_min[b]}, {x_max[b]}, {y_min[b]}, {y_min[b]}, {z_min[b]}, {z_max[b]}, RADIUS = 0.225, XYZ = {fan_table[r][0]:.4f}, {y_min[b]}, {fan_table[r][2]:.4f}/")
              print(f"&HVAC ID = 'N{b}01', TYPE_ID = 'NODE', DUCT_ID = 'D{b}', VENT_ID = 'V{b}01'/")
              print(f"&HVAC ID = 'N{b}02', TYPE_ID = 'NODE', DUCT_ID = 'D{b}', VENT_ID = 'V{b}02'/")
              print(f"&HVAC ID = 'D{b}', TYPE_ID = 'DUCT', DIAMETER = 0.45, VOLUME_FLOW = {volumeflow:.4f}, NODE_ID = 'N{b}01', 'N{b}02', ROUGHNESS = 1.0E-3, LENGTH = {fan_length:.4f}/")
              b += 1

          # contain uvw
          else:
              print(f"&VENT ID = 'V{b}01', SURF_ID = 'HVAC', XB = {x_min[b]}, {x_max[b]}, {y_max[b]}, {y_max[b]}, {z_min[b]}, {z_max[b]}, RADIUS = 0.225, XYZ = {fan_table[r][0]:.4f}, {y_max[b]}, {fan_table[r][2]:.4f}/")
              print(f"&VENT ID = 'V{b}02', SURF_ID = 'HVAC', XB = {x_min[b]}, {x_max[b]}, {y_min[b]}, {y_min[b]}, {z_min[b]}, {z_max[b]}, RADIUS = 0.225, XYZ = {fan_table[r][0]:.4f}, {y_min[b]}, {fan_table[r][2]:.4f}, UVW={fan_table[r][4]:.4f},{fan_table[r][5]:.4f},{fan_table[r][6]:.4f}/")
              print(f"&HVAC ID = 'N{b}01', TYPE_ID = 'NODE', DUCT_ID = 'D{b}', VENT_ID = 'V{b}01'/")
              print(f"&HVAC ID = 'N{b}02', TYPE_ID = 'NODE', DUCT_ID = 'D{b}', VENT_ID = 'V{b}02'/")
              print(f"&HVAC ID = 'D{b}', TYPE_ID = 'DUCT', DIAMETER = 0.45, VOLUME_FLOW = {volumeflow:.4f}, NODE_ID = 'N{b}01', 'N{b}02', ROUGHNESS = 1.0E-3, LENGTH = {fan_length:.4f}/")
              b += 1


